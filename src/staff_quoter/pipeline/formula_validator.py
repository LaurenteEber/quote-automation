from __future__ import annotations

from pathlib import Path
import re

from openpyxl import load_workbook

from .models import FormulaIssue, FormulaValidationReport


class WorkbookFormulaValidator:
    BAD_TOKENS = ["#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?"]

    _SHEET_REF_PATTERN = re.compile(r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!")
    _STRING_LITERAL_PATTERN = re.compile(r'"(?:[^"]|"")*"')

    def validate(self, workbook_path: Path | str) -> FormulaValidationReport:
        workbook_path = Path(workbook_path)
        wb = load_workbook(workbook_path, data_only=False, read_only=True)

        known_sheets = set(wb.sheetnames)
        total_formulas = 0
        issues: list[FormulaIssue] = []

        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if cell.data_type != "f":
                        continue
                    if not isinstance(value, str):
                        continue

                    total_formulas += 1
                    upper_formula = value.upper()

                    for token in self.BAD_TOKENS:
                        if token in upper_formula:
                            issues.append(
                                FormulaIssue(
                                    sheet=ws.title,
                                    cell=cell.coordinate,
                                    code="ERROR_TOKEN",
                                    detail=f"Found token {token}",
                                )
                            )

                    formula_without_strings = self._STRING_LITERAL_PATTERN.sub("", value)
                    for match in self._SHEET_REF_PATTERN.finditer(formula_without_strings):
                        ref_sheet = (match.group(1) or match.group(2) or "").strip()
                        if ref_sheet and ref_sheet not in known_sheets:
                            issues.append(
                                FormulaIssue(
                                    sheet=ws.title,
                                    cell=cell.coordinate,
                                    code="UNKNOWN_SHEET_REF",
                                    detail=f"Unknown sheet reference: {ref_sheet}",
                                )
                            )

        wb.close()
        return FormulaValidationReport(
            workbook_path=str(workbook_path),
            total_formulas=total_formulas,
            issues=_dedupe_issues(issues),
        )


def _dedupe_issues(issues: list[FormulaIssue]) -> list[FormulaIssue]:
    unique: dict[tuple[str, str, str, str], FormulaIssue] = {}
    for issue in issues:
        key = (issue.sheet, issue.cell, issue.code, issue.detail)
        unique[key] = issue
    return list(unique.values())
