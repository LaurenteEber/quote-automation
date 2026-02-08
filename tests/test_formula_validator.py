from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from staff_quoter.pipeline import WorkbookFormulaValidator


def test_formula_validator_detects_tokens_and_unknown_sheets(tmp_path: Path) -> None:
    workbook_path = tmp_path / "validator_case.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "MAIN"
    wb.create_sheet("DATA")

    ws["A1"] = "=1+1"
    ws["A2"] = "=#REF!"
    ws["A3"] = "=MISSING_SHEET!A1"

    wb.save(workbook_path)
    wb.close()

    report = WorkbookFormulaValidator().validate(workbook_path)

    assert report.total_formulas == 3
    assert report.has_errors
    codes = {issue.code for issue in report.issues}
    assert "ERROR_TOKEN" in codes
    assert "UNKNOWN_SHEET_REF" in codes
