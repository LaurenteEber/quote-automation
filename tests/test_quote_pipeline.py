from __future__ import annotations

from pathlib import Path
import json

from openpyxl import Workbook

from staff_quoter.config import Settings
from staff_quoter.pipeline import QuotePipeline


def _create_workbook(path: Path) -> None:
    wb = Workbook()

    ws_input = wb.active
    ws_input.title = "INPUT_QUOTE"
    ws_input["A2"] = "Q-TEST-001"
    ws_input["C2"] = "MACHINING"
    ws_input["D2"] = "Test Customer"
    ws_input["G2"] = "PN-TEST-001"

    ws_calc = wb.create_sheet("CALC_OUTPUTS")
    ws_calc["B2"] = 100.0
    ws_calc["C2"] = 150.0
    ws_calc["D2"] = 0.5
    ws_calc["E2"] = 2.0
    ws_calc["F2"] = 50

    ws_quote = wb.create_sheet("QUOTE_OUTPUT")
    ws_quote["C2"] = "TRUE"

    wb.save(path)
    wb.close()


def test_quote_pipeline_generates_json_and_pdf(tmp_path: Path) -> None:
    workbook_path = tmp_path / "pipeline_case.xlsx"
    _create_workbook(workbook_path)

    settings = Settings(
        workspace_root=tmp_path,
        google_credentials_file="",
        google_sheets_id="",
        xlsx_recalc_script=tmp_path / "recalc.py",
        default_workbook=workbook_path,
        output_json_dir=tmp_path / "output" / "json",
        output_pdf_dir=tmp_path / "output" / "pdf",
    )

    result = QuotePipeline(settings).run(workbook_path, fail_on_formula_issues=True, run_recalc=False)

    json_path = Path(result.json_output_path)
    pdf_path = Path(result.pdf_output_path)

    assert json_path.exists()
    assert pdf_path.exists()
    assert pdf_path.stat().st_size > 0

    payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert payload["quote_id"] == "Q-TEST-001"
    assert payload["engine_type"] == "MACHINING"
    assert payload["total_price"] == 150.0
    assert payload["pdf_ready_flag"] is True

    assert result.formula_report["total_formulas"] == 0
    assert result.formula_report["issue_count"] == 0
